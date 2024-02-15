"""
Name: Import GRID layout 
Autor: Andrea Sulova
Date: 14th Feb 2024

"""
import arcpy
import openpyxl
import os
import pandas as pd
import zipfile
from arcpy import metadata as md

#------------ Inputs
# Define input parameters fetched from the user or other sources

file_path = arcpy.GetParameterAsText(0)

spatial_reference = arcpy.GetParameterAsText(1)

output_gdb = arcpy.GetParameterAsText(2)

output_fc = arcpy.GetParameterAsText(3)

data_inventory = arcpy.GetParameterAsText(4)

sheet_data_inventory = arcpy.GetParameterAsText(5)

arcpy.env.workspace = output_gdb

# -------- Import excel files to Dataframe

# Beate's amazing work for importing table from the IAC template provided by Cable Engineers
def find_table_in_excel(file_path, keyword):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Loop through each sheet in the workbook
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Loop through each row in the sheet
        for row in ws.iter_rows():
            # Loop through each cell in the row
            for cell in row:
                # Check if the cell value matches the keyword
                if cell.value == keyword:
                    start_cell = cell.coordinate

                    # Get start cell coordinates
                    start_col = start_cell[0]
                    start_row = int(start_cell[1:])

                    # Initialize end cell coordinates
                    end_col = start_col
                    end_row = start_row

                    # Find end of table in rows
                    while ws[end_col + str(end_row)].value is not None:
                        end_row += 1

                    # Find end of table in columns
                    while ws[end_col + str(start_row)].value is not None:
                        if ord(end_col) < 90:  # ASCII value of 'Z'
                            end_col = chr(ord(end_col) + 1)
                        else:
                            break

                    # Adjust for overstepping boundaries
                    if ord(end_col) > 65:  # ASCII value of 'A'
                        end_col = chr(ord(end_col) - 1)
                        
                    end_cell = end_col + str(end_row - 1)
                    
                    # Return table range
                    return sheet, start_cell, end_cell                  
                    
def excel_table_to_feature_class(excel_file_path, output_gdb_path, output_fc, sheet_name, start_cell, end_cell, epsg_code):
    
    epsg_code = int(epsg_code)
    
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb[sheet_name]

    # Create an empty list to store the cable strings and a dictionary to store the attributes
    cable_strings = []
    attributes = {}
    string_number = 0

    # Get the headers from the first row of the cell range
    headers = [cell.value for cell in ws[start_cell:end_cell][0]]
    
    # As the coordinate columns for start and end point have same names, they need to get renamed to be unique
    replacements = {'Easting [m]': ['Start Easting [m]', 'End Easting [m]'], 
                    'Northing [m]': ['Start Northing [m]', 'End Northing [m]'],
                    'Depth to LAT [m]': ['Start Depth to LAT [m]', 'End Depth to LAT [m]']}
    for key, values in replacements.items():
        try:
            for value in values:
                index = headers.index(key)
                headers[index] = value
        except ValueError:
            pass
        
    arcpy.AddMessage(f'Header row from Excel: {headers}')

    # Create a dictionary to map valid field names to original headers
    field_names = {arcpy.ValidateFieldName(header): header for header in headers}
    field_names = {name.rstrip("_").replace("__","_"): value for name, value in field_names.items()}
    
    # Iterate through each row in the cell range (excluding the header row)
    for row in ws[start_cell:end_cell][1:]:
        # Create a polyline geometry object using the start and end coordinates
        array = arcpy.Array([arcpy.Point(row[1].value, row[2].value), arcpy.Point(row[5].value, row[6].value)])
        polyline = arcpy.Polyline(array)
    
        # Check if the start point is 0 and increment string number
        if row[0].value == 0:
            string_number += 1
    
        # Create a dictionary of attributes for the current row
        attrs = {'String_number' : string_number}
        attrs.update({header: value for header, value in zip(field_names, row)})
    
        # Append a tuple with the polyline and the attributes to the list
        cable_strings.append((polyline, attrs))

    # Define the spatial reference of the output feature class (you may need to adjust this)
    spatial_reference = arcpy.SpatialReference(epsg_code)
    arcpy.AddMessage(f'Create feature class: {output_fc}')

    # Create a new feature class in the geodatabase
    fc_path = os.path.join(output_gdb_path, output_fc)
    arcpy.CreateFeatureclass_management(output_gdb_path, output_fc, 'POLYLINE', spatial_reference=spatial_reference)

    # String number should be the first attribute in fc table, so an extra step is needed to get it added to field_names dict
    string_item = ({"String_number" : "String number"})
    field_names = {**string_item, **field_names}

    # Add fields for each header in the Excel table
    # arcpy.AddField_management(fc_path, 'string_number', 'LONG', field_alias="String number")
    for field_name, header in field_names.items():
        if header in ['Start point', 'End point', "String number"]:
            arcpy.AddField_management(fc_path, field_name, 'SHORT', field_alias=header)
        else:
            arcpy.AddField_management(fc_path, field_name, 'DOUBLE', field_alias=header)

    # Use an insert cursor to add the cable strings to the new feature class
    arcpy.AddMessage(f'Write cable strings to fc')
    
    try:
        with arcpy.da.InsertCursor(fc_path, ['SHAPE@'] + list(field_names.keys())) as cursor:
            for cable_string, attrs in cable_strings:
                cursor.insertRow([cable_string] + [attrs[field_name].value if hasattr(attrs[field_name], 'value') else attrs[field_name] for field_name in field_names])
    except Exception as e:
        arcpy.AddError(f"An error occurred: {e}")
    arcpy.AddMessage(f"*** Finished ***")
   

# Call function to get the sheet name and cell range
sheet_name, start_cell, end_cell = find_table_in_excel(file_path, "Start point")

# Create a new feature class in the geodatabase
fc_path = os.path.join(output_gdb, output_fc)

if not arcpy.Exists(fc_path):
    excel_table_to_feature_class(file_path, output_gdb, output_fc, sheet_name, start_cell, end_cell, spatial_reference) 
    arcpy.AddMessage("The feature class in geodatabase is created successfully")
else:
    arcpy.AddMessage("The feature class already exists in geodatabase")

#--------- Metadata extracted from the data inventory file

# Read dataframe = excel file
df = pd.read_excel(data_inventory,sheet_data_inventory)

# Filter the DataFrame based on the 'Full Name' column containing the search string
filtered_df = df[df["Full Name"] == output_fc]
arcpy.AddMessage(filtered_df)

# Check if the search string was found in the DataFrame

if not filtered_df.empty:
    # Get the text from text_column in the same row
    imported_title = filtered_df.iloc[0]["Full Name"]
    imported_summary = filtered_df.iloc[0]["Summary"]
    imported_tags = filtered_df.iloc[0]["Tags"]
    imported_Description = filtered_df.iloc[0]["Description"]
    imported_Credits = filtered_df.iloc[0]["Credits"]
    imported_Date = str(filtered_df.iloc[0]["Date Created"])
    imported_Description = imported_Description + "\n" + imported_Date
    
    new_md = md.Metadata()
    new_md.title = imported_title
    new_md.tags = imported_tags
    new_md.summary = imported_summary
    new_md.description = imported_Description
    new_md.credits = imported_Credits

    # Assign the Metadata object's content to a target item
    tgt_item_md = md.Metadata(fc_path)

    if not tgt_item_md.isReadOnly:
        tgt_item_md.copy(new_md)
        tgt_item_md.save()

    arcpy.AddMessage("Adding metadata is completed successfully")
    
else:
    arcpy.arcpy.AddMessage("Please add information into the DATA INVENTORY File (Excel)")
    arcpy.arcpy.AddMessage("Name of feature class should be the same as in the data inventory (column Name)")


#-------  Establishing Alias Names
    
# Get the describe object for the feature class
desc = arcpy.Describe(fc_path)
current_alias = desc.aliasName
arcpy.AddMessage(current_alias)

# Create a new alias name by replacing underscores with spaces
new_alias = output_fc.replace("_", " ")

try:
    # Set the new alias name for the feature class
    arcpy.AlterAliasName(fc_path, new_alias)
    arcpy.AddMessage("Alias name changed successfully.")

except Exception as e:
    arcpy.AddMessage("An error occurred during creating alias name")


#------  Export this feature class to Shapefile
    
# Create the full path to the folder where the shapefile will be save

directory = os.path.dirname(file_path)
shp_folder = os.path.join(directory, output_fc)
shp_file = output_fc + ".shp"
output_shp = os.path.join(shp_folder,shp_file)
arcpy.AddMessage(output_shp)

# Check if the folder already exists before creating it
if not os.path.exists(shp_folder):
    # If the folder does not exist, create it and convert the feature class to a shapefile
    os.makedirs(shp_folder)
    arcpy.FeatureClassToFeatureClass_conversion(fc_path, shp_folder, shp_file)
    arcpy.AddMessage("Shapefile is created successfully and saved in this folder:")
else:
    # If the folder already exists, inform the user
    arcpy.AddMessage("SHP Folder already exists.")


#------ Feature class to a DWG file for CAD
    
# Set up paths and file names
dwg_folder = os.path.join(shp_folder+ '_DWG')

# Output DWG file
dwg_output = os.path.join(dwg_folder, output_fc +'.dwg')

# Check if the output path exists, if not, create it
if not os.path.exists(dwg_folder):
    os.makedirs(dwg_folder)
    arcpy.conversion.ExportCAD(output_shp, "DWG_R2018", dwg_output, False, False)
    arcpy.AddMessage("DWG file is created successfully and saved in this folder:")
else:
    arcpy.AddMessage("DWG Folder already exists.")
